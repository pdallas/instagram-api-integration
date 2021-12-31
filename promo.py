# Run this code to get the latest 25 post insights for each instagram account
from openpyxl import load_workbook
from defines import getCreds, makeApiCall
from post_insights import getUserMedia, getMediaInsights
from get_insta_acc import getIgUserInfo, getIGid

params = getCreds()  # get creds
params['debug'] = 'no'
response = getIGid(params)

for cnt, igids in enumerate(response):
    # load workbook
    wb = load_workbook(filename='insights.xlsx')
    dest_filename = 'insights.xlsx'
    ws = wb.create_sheet(title=response[cnt])
    params = getCreds()
    params['debug'] = 'no'
    params["instagram_account_id"] = response[cnt]
    resp = getUserMedia(params)  # get users last 25 posts from the api for each account id
    i = 0
    for count, r in enumerate(resp['json_data']['data']):

        params['latest_media_id'] = resp['json_data']['data'][count]['id']  # store latest post id
        if 'VIDEO' == resp['json_data']['data'][count]['media_type']:  # media is a video
            params['metric'] = 'engagement,impressions,reach,saved,video_views'
        else:  # media is an image
            params['metric'] = 'engagement,impressions,reach,saved'

        media_insights = getMediaInsights(params)
        # A COLUMN
        ws['A' + str(i + 1)] = 'POST NUMBER'
        ws['A' + str(i + 2)] = 'LINK TO POST:'
        ws['A' + str(i + 3)] = 'DATE POSTED:'
        ws['A' + str(i + 4)] = 'MEDIA TYPE:'
        ws['A' + str(i + 5)] = 'ENGAGEMENT(LIFETIME):'
        ws['A' + str(i + 6)] = 'IMPRESSIONS(LIFETIME):'
        ws['A' + str(i + 7)] = 'REACH(LIFETIME):'
        ws['A' + str(i + 8)] = 'SAVED(LIFETIME):'
        ws['A' + str(i + 9)] = 'VIDEO VIEWS(LIFETIME):'
        ws['A' + str(i + 10)] = 'POST CAPTION:'
        # B COLUMN
        ws['B' + str(i + 1)] = count + 1
        ws['B' + str(i + 2)] = resp['json_data']['data'][count]['permalink']
        ws['B' + str(i + 3)] = resp['json_data']['data'][count]['timestamp']
        ws['B' + str(i + 4)] = resp['json_data']['data'][count]['media_type']

        print("\n---- POST INSIGHTS -----\n")  # section header

        for counter, insight in enumerate(media_insights['json_data']['data']):  # loop over post insights
            if counter == 0:
                ws['B' + str(i + 5)] = insight['values'][0]['value']
            elif counter == 1:
                ws['B' + str(i + 6)] = insight['values'][0]['value']
            elif counter == 2:
                ws['B' + str(i + 7)] = insight['values'][0]['value']
            elif counter == 3:
                ws['B' + str(i + 8)] = insight['values'][0]['value']
            elif counter == 4:
                ws['B' + str(i + 9)] = insight['values'][0]['value']
            print("\t" + insight['title'] + " (" + insight['period'] + "): " + str(insight['values'][0]['value']))  # display info

        ws['B' + str(i + 10)] = resp['json_data']['data'][count]['caption']
        i = i + 12

        print("\n---- " + str(resp['json_data']['data'][count]['username']) +" POST num "+ str(count+1) + " -----\n")
        print("\tLink to post:")  # link to post
        print("\t" + resp['json_data']['data'][count]['permalink'])  # link to post
        print("\n\tPost caption:")  # post caption
        print("\t" + resp['json_data']['data'][count]['caption'])  # post caption
        print("\n\tMedia Type:")  # type of media
        print("\t" + resp['json_data']['data'][count]['media_type'])  # type of media
        print("\n\tPosted at:")  # when it was posted
        print("\t" + resp['json_data']['data'][count]['timestamp'])  # when it was posted
    wb.save(filename=dest_filename)
